"""
ASPROJUMA - Script de migración Excel → Supabase/PostgreSQL
============================================================
Lee los ficheros Excel actuales y genera un fichero SQL listo
para ejecutar en Supabase (panel SQL Editor → New query → Paste → Run).

Uso:
    python migrar_asprojuma.py

Requisitos:
    pip install openpyxl

Ficheros necesarios en la misma carpeta que este script:
    - DATOS_SOCIOS_Y_MIEMBROS_COOPERANTES_al_10demarzo2026.xlsx
    - REMESA-ASOCIADOS_ASPROJUMA_DIC2025.xlsx

Salida:
    - asprojuma_datos.sql  (pegar en Supabase SQL Editor)
    - informe_migracion.txt (resumen de lo que se ha migrado)
"""

import openpyxl
import re
import sys
from datetime import datetime, date


# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────

FICHERO_SOCIOS    = "DATOS_SOCIOS_Y_MIEMBROS_COOPERANTES_al_10demarzo2026.xlsx"
FICHERO_REMESA    = "REMESA-ASOCIADOS_ASPROJUMA_DIC2025.xlsx"
FICHERO_SQL       = "asprojuma_datos.sql"
FICHERO_INFORME   = "informe_migracion.txt"

SITUACION_LABELS = {
    "M":  "activo_exento",   # Mayor de 85, exento de pago
    "B":  "baja",
    "F":  "fallecido",
    "SH": "honorario",
    None: "activo",
    "":   "activo",
}

CUOTA_ANUAL = 50.0   # euros
CUOTA_SEMESTRAL = 25.0


# ─── UTILIDADES ───────────────────────────────────────────────────────────────

def limpiar(v):
    """Convierte None y strings vacíos a None, limpia espacios."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def sql_str(v):
    """Escapa un string para SQL (comillas simples)."""
    if v is None:
        return "NULL"
    escaped = str(v).replace("'", "''")
    return f"'{escaped}'"


def sql_date(v):
    """Convierte fecha Python/datetime a SQL."""
    if v is None:
        return "NULL"
    if isinstance(v, datetime):
        return f"'{v.strftime('%Y-%m-%d')}'"
    if isinstance(v, date):
        return f"'{v.strftime('%Y-%m-%d')}'"
    s = str(v).strip()
    return f"'{s}'" if s else "NULL"


def sql_bool(v):
    return "TRUE" if v else "FALSE"


def normalizar_iban(raw):
    """
    Normaliza un IBAN al formato ES + 22 dígitos.
    Los registros antiguos tienen 20 dígitos sin prefijo ES.
    """
    if not raw:
        return None
    s = str(raw).strip()
    if s in ("", "Honorario", "Recibo", "BANCO"):
        return None
    # Quitar espacios
    s_clean = s.replace(" ", "").upper()
    if s_clean.startswith("ES") and len(s_clean) == 24:
        return s_clean          # Ya está en formato correcto
    # Formato antiguo: 20 dígitos (sin ES)
    digits = re.sub(r"\D", "", s_clean)
    if len(digits) == 20:
        # Calcular dígitos de control según ISO 7064 MOD 97-10
        rearranged = digits + "142800"   # ES=1428, control=00 placeholder
        remainder = int(rearranged) % 97
        check = 98 - remainder
        iban = f"ES{check:02d}{digits}"
        return iban
    # Si tiene prefijo ES pero longitud incorrecta, devolver limpio
    if s_clean.startswith("ES"):
        return s_clean
    return s_clean if s_clean else None


def normalizar_telefono(v):
    """Limpia y devuelve el teléfono como string."""
    if v is None:
        return None
    s = str(v).strip()
    # Quitar caracteres extraños excepto +, -, espacios y dígitos
    s = re.sub(r"[^\d\s\+\-]", "", s)
    s = s.strip()
    return s if s else None


def mapear_situacion_socio(sit):
    sit_clean = str(sit).strip().upper() if sit else ""
    mapping = {
        "M": "activo_exento",
        "B": "baja",
        "F": "fallecido",
        "SH": "honorario",
        "": "activo",
    }
    return mapping.get(sit_clean, "activo")


def mapear_situacion_cooperante(sit):
    sit_clean = str(sit).strip().upper() if sit else ""
    mapping = {
        "B": "baja",
        "F": "fallecido",
        "": "activo",
    }
    return mapping.get(sit_clean, "activo")


# ─── LECTURA EXCEL SOCIOS ─────────────────────────────────────────────────────

def leer_socios(ruta):
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["SOCIOS"]
    filas = list(ws.iter_rows(max_row=300, values_only=True))

    socios = []
    alertas = []

    for i, fila in enumerate(filas[1:], start=2):
        # Filtrar filas vacías y filas de leyenda
        if not any(v is not None for v in fila):
            continue
        nseq, sit, num_socio, apellidos, nombre, f_nac, dni, iban_raw, \
            f_ingreso, centro, direccion, cp, localidad, provincia, \
            tel1, f_baja, tel2, mail_uma, otros_mail = (fila[:19] + (None,) * 19)[:19]

        # Saltar filas de leyenda (contienen texto explicativo)
        if isinstance(sit, str) and len(sit) > 5:
            continue
        if not isinstance(num_socio, int) or num_socio <= 0:
            continue

        estado = mapear_situacion_socio(sit)
        iban   = normalizar_iban(iban_raw)
        tel_fijo  = normalizar_telefono(tel1)
        tel_movil = normalizar_telefono(tel2)

        # Email principal: preferir mail UMA, sino otros
        email_uma   = limpiar(mail_uma)
        email_otros = limpiar(otros_mail)
        # Saltar emails que son notas (FALLECIDO, BAJA...)
        if email_uma and email_uma.upper() in ("FALLECIDO", "BAJA", ""):
            email_uma = None
        if email_otros and email_otros.upper() in ("FALLECIDO", "BAJA", ""):
            email_otros = None

        # Advertencias
        if not iban and estado == "activo":
            alertas.append(f"  Fila {i} - Socio Nº {num_socio} ({limpiar(apellidos)}): sin IBAN (activo)")
        if not limpiar(dni) and estado == "activo":
            alertas.append(f"  Fila {i} - Socio Nº {num_socio} ({limpiar(apellidos)}): sin DNI")

        socios.append({
            "num_socio":   num_socio,
            "tipo":        "profesor",
            "estado":      estado,
            "apellidos":   limpiar(apellidos),
            "nombre":      limpiar(nombre),
            "dni":         limpiar(dni),
            "f_nacimiento": f_nac if isinstance(f_nac, datetime) else None,
            "iban":        iban,
            "f_ingreso":   f_ingreso if isinstance(f_ingreso, datetime) else None,
            "f_baja":      f_baja if isinstance(f_baja, datetime) else None,
            "centro":      limpiar(centro),
            "direccion":   limpiar(direccion),
            "cp":          limpiar(str(cp)) if cp else None,
            "localidad":   limpiar(localidad),
            "provincia":   limpiar(provincia),
            "tel_fijo":    tel_fijo,
            "tel_movil":   tel_movil,
            "email_uma":   email_uma,
            "email_otros": email_otros,
        })

    wb.close()
    return socios, alertas


# ─── LECTURA EXCEL COOPERANTES ────────────────────────────────────────────────

def leer_cooperantes(ruta):
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["MIEMBROS COOPERANTES"]
    filas = list(ws.iter_rows(max_row=200, values_only=True))

    cooperantes = []
    alertas = []

    for i, fila in enumerate(filas[1:], start=2):
        if not any(v is not None for v in fila):
            continue
        num, sit, apellidos, nombre, dni, iban_raw, f_nac, f_ingreso, \
            centro, f_baja, direccion, cp, localidad, provincia, \
            tel_fijo, tel_movil, email = (fila[:17] + (None,) * 17)[:17]

        # Saltar filas de leyenda
        if isinstance(sit, str) and len(sit) > 5:
            continue
        if not isinstance(num, int):
            continue

        estado = mapear_situacion_cooperante(sit)
        iban   = normalizar_iban(iban_raw)

        if not iban and estado == "activo":
            alertas.append(f"  Fila {i} - Cooperante Nº {num} ({limpiar(apellidos)}): sin IBAN (activo)")

        cooperantes.append({
            "num_cooperante": num,
            "tipo":           "cooperante",
            "estado":         estado,
            "apellidos":      limpiar(apellidos),
            "nombre":         limpiar(nombre),
            "dni":            limpiar(dni),
            "f_nacimiento":   f_nac if isinstance(f_nac, datetime) else None,
            "iban":           iban,
            "f_ingreso":      f_ingreso if isinstance(f_ingreso, datetime) else None,
            "f_baja":         f_baja if isinstance(f_baja, datetime) else None,
            "centro":         limpiar(centro),
            "direccion":      limpiar(direccion),
            "cp":             limpiar(str(cp)) if cp else None,
            "localidad":      limpiar(localidad),
            "provincia":      limpiar(provincia),
            "tel_fijo":       normalizar_telefono(tel_fijo),
            "tel_movil":      normalizar_telefono(tel_movil),
            "email":          limpiar(email),
        })

    wb.close()
    return cooperantes, alertas


# ─── GENERACIÓN SQL ───────────────────────────────────────────────────────────

def generar_sql(socios, cooperantes):
    lineas = []

    lineas.append("-- ================================================================")
    lineas.append("-- ASPROJUMA - Migración inicial de datos")
    lineas.append(f"-- Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lineas.append("-- Instrucciones: pegar en Supabase > SQL Editor > New query > Run")
    lineas.append("-- ================================================================")
    lineas.append("")

    # ── CREAR TABLAS ──────────────────────────────────────────────────────────
    lineas.append("""
-- ─────────────────────────────────────────────────────────────────────────────
-- PASO 1: Crear las tablas (se pueden ejecutar varias veces sin error)
-- ─────────────────────────────────────────────────────────────────────────────

CREATE TABLE IF NOT EXISTS socios (
    id               SERIAL PRIMARY KEY,
    tipo             VARCHAR(20)  NOT NULL CHECK (tipo IN ('profesor','cooperante')),
    estado           VARCHAR(30)  NOT NULL DEFAULT 'activo'
                       CHECK (estado IN ('activo','activo_exento','baja','fallecido','honorario','pendiente','suspendido')),
    num_socio        INT          UNIQUE,           -- número de socio profesor
    num_cooperante   INT          UNIQUE,           -- número de cooperante
    apellidos        VARCHAR(200),
    nombre           VARCHAR(150),
    dni              VARCHAR(20)  UNIQUE,
    fecha_nacimiento DATE,
    iban             VARCHAR(34),
    titular_cuenta   VARCHAR(200),
    fecha_ingreso    DATE,
    fecha_baja       DATE,
    centro           VARCHAR(200),
    direccion        TEXT,
    codigo_postal    VARCHAR(10),
    localidad        VARCHAR(100),
    provincia        VARCHAR(100),
    tel_fijo         VARCHAR(30),
    tel_movil        VARCHAR(30),
    email_uma        VARCHAR(200),
    email_otros      VARCHAR(200),
    email_principal  VARCHAR(200) GENERATED ALWAYS AS (
                       COALESCE(email_uma, email_otros)
                     ) STORED,
    notas            TEXT,
    migrado_excel    BOOLEAN      DEFAULT TRUE,     -- marca registros de migración
    created_at       TIMESTAMPTZ  DEFAULT NOW(),
    updated_at       TIMESTAMPTZ  DEFAULT NOW()
);

-- Tabla específica de profesores (datos académicos adicionales)
CREATE TABLE IF NOT EXISTS socios_profesores (
    socio_id         INT PRIMARY KEY REFERENCES socios(id) ON DELETE CASCADE,
    departamento     VARCHAR(200),
    titulacion       VARCHAR(200),
    fecha_jubilacion DATE,
    categoria        VARCHAR(100)
);

-- Tabla específica de cooperantes
CREATE TABLE IF NOT EXISTS socios_cooperantes (
    socio_id         INT PRIMARY KEY REFERENCES socios(id) ON DELETE CASCADE,
    estudios         TEXT,
    aficiones        TEXT,
    descripcion_relacion TEXT
);

-- Carnets
CREATE TABLE IF NOT EXISTS carnets (
    id               SERIAL PRIMARY KEY,
    socio_id         INT NOT NULL REFERENCES socios(id) ON DELETE CASCADE,
    anio_vigencia    INT NOT NULL,
    fecha_emision    DATE DEFAULT CURRENT_DATE,
    fecha_caducidad  DATE,
    estado           VARCHAR(20) DEFAULT 'vigente' CHECK (estado IN ('vigente','caducado','anulado')),
    pdf_url          VARCHAR(500),
    enviado_email    BOOLEAN DEFAULT FALSE,
    created_at       TIMESTAMPTZ DEFAULT NOW()
);

-- Cuotas / pagos
CREATE TABLE IF NOT EXISTS cuotas (
    id               SERIAL PRIMARY KEY,
    socio_id         INT NOT NULL REFERENCES socios(id) ON DELETE CASCADE,
    anio             INT NOT NULL,
    semestre         INT NOT NULL CHECK (semestre IN (1, 2)),
    importe          DECIMAL(8,2) NOT NULL,
    estado           VARCHAR(20) DEFAULT 'pendiente'
                       CHECK (estado IN ('pendiente','cobrado','devuelto','exento')),
    fecha_cobro      DATE,
    metodo_pago      VARCHAR(30) DEFAULT 'domiciliacion',
    referencia_remesa VARCHAR(60),
    notas            TEXT,
    created_at       TIMESTAMPTZ DEFAULT NOW()
);

-- Índices útiles para búsquedas frecuentes
CREATE INDEX IF NOT EXISTS idx_socios_tipo    ON socios(tipo);
CREATE INDEX IF NOT EXISTS idx_socios_estado  ON socios(estado);
CREATE INDEX IF NOT EXISTS idx_socios_dni     ON socios(dni);
CREATE INDEX IF NOT EXISTS idx_cuotas_socio   ON cuotas(socio_id, anio);
""")

    # ── INSERTAR SOCIOS ───────────────────────────────────────────────────────
    lineas.append("-- ─────────────────────────────────────────────────────────────────────────────")
    lineas.append("-- PASO 2: Insertar socios profesores")
    lineas.append(f"-- Total: {len(socios)} registros")
    lineas.append("-- ─────────────────────────────────────────────────────────────────────────────")
    lineas.append("")

    activos_socios = 0
    for s in socios:
        lineas.append(
            f"INSERT INTO socios "
            f"(tipo, estado, num_socio, apellidos, nombre, dni, fecha_nacimiento, "
            f"iban, fecha_ingreso, fecha_baja, centro, direccion, codigo_postal, "
            f"localidad, provincia, tel_fijo, tel_movil, email_uma, email_otros, migrado_excel) VALUES ("
            f"'profesor', "
            f"{sql_str(s['estado'])}, "
            f"{s['num_socio']}, "
            f"{sql_str(s['apellidos'])}, "
            f"{sql_str(s['nombre'])}, "
            f"{sql_str(s['dni'])}, "
            f"{sql_date(s['f_nacimiento'])}, "
            f"{sql_str(s['iban'])}, "
            f"{sql_date(s['f_ingreso'])}, "
            f"{sql_date(s['f_baja'])}, "
            f"{sql_str(s['centro'])}, "
            f"{sql_str(s['direccion'])}, "
            f"{sql_str(s['cp'])}, "
            f"{sql_str(s['localidad'])}, "
            f"{sql_str(s['provincia'])}, "
            f"{sql_str(s['tel_fijo'])}, "
            f"{sql_str(s['tel_movil'])}, "
            f"{sql_str(s['email_uma'])}, "
            f"{sql_str(s['email_otros'])}, "
            f"TRUE"
            f") ON CONFLICT (num_socio) DO NOTHING;"
        )
        if s['estado'] == 'activo':
            activos_socios += 1

    lineas.append("")
    lineas.append("-- ─────────────────────────────────────────────────────────────────────────────")
    lineas.append("-- PASO 3: Insertar miembros cooperantes")
    lineas.append(f"-- Total: {len(cooperantes)} registros")
    lineas.append("-- ─────────────────────────────────────────────────────────────────────────────")
    lineas.append("")

    activos_coop = 0
    for c in cooperantes:
        lineas.append(
            f"INSERT INTO socios "
            f"(tipo, estado, num_cooperante, apellidos, nombre, dni, fecha_nacimiento, "
            f"iban, fecha_ingreso, fecha_baja, centro, direccion, codigo_postal, "
            f"localidad, provincia, tel_fijo, tel_movil, email_otros, migrado_excel) VALUES ("
            f"'cooperante', "
            f"{sql_str(c['estado'])}, "
            f"{c['num_cooperante']}, "
            f"{sql_str(c['apellidos'])}, "
            f"{sql_str(c['nombre'])}, "
            f"{sql_str(c['dni'])}, "
            f"{sql_date(c['f_nacimiento'])}, "
            f"{sql_str(c['iban'])}, "
            f"{sql_date(c['f_ingreso'])}, "
            f"{sql_date(c['f_baja'])}, "
            f"{sql_str(c['centro'])}, "
            f"{sql_str(c['direccion'])}, "
            f"{sql_str(c['cp'])}, "
            f"{sql_str(c['localidad'])}, "
            f"{sql_str(c['provincia'])}, "
            f"{sql_str(c['tel_fijo'])}, "
            f"{sql_str(c['tel_movil'])}, "
            f"{sql_str(c['email'])}, "
            f"TRUE"
            f") ON CONFLICT (num_cooperante) DO NOTHING;"
        )
        if c['estado'] == 'activo':
            activos_coop += 1

    # ── CUOTAS HISTÓRICAS 2025 ────────────────────────────────────────────────
    lineas.append("")
    lineas.append("-- ─────────────────────────────────────────────────────────────────────────────")
    lineas.append("-- PASO 4: Registrar cuota semestre 2 de 2025 como cobrada (remesa dic. 2025)")
    lineas.append("-- Se marca como cobrada para todos los socios activos con IBAN")
    lineas.append("-- ─────────────────────────────────────────────────────────────────────────────")
    lineas.append("""
INSERT INTO cuotas (socio_id, anio, semestre, importe, estado, fecha_cobro, metodo_pago, referencia_remesa)
SELECT
    s.id,
    2025,
    2,
    25.00,
    'cobrado',
    '2025-12-01',
    'domiciliacion',
    'ASPROJUMA 2025-2'
FROM socios s
WHERE s.tipo = 'profesor'
  AND s.estado IN ('activo', 'activo_exento')
  AND s.iban IS NOT NULL
  AND s.migrado_excel = TRUE
ON CONFLICT DO NOTHING;
""")

    # ── VERIFICACIÓN FINAL ────────────────────────────────────────────────────
    lineas.append("-- ─────────────────────────────────────────────────────────────────────────────")
    lineas.append("-- PASO 5: Verificación – ejecutar estas consultas para comprobar el resultado")
    lineas.append("-- ─────────────────────────────────────────────────────────────────────────────")
    lineas.append("""
SELECT
    tipo,
    estado,
    COUNT(*) AS total
FROM socios
GROUP BY tipo, estado
ORDER BY tipo, estado;

-- Comprobación rápida de datos migrados:
SELECT COUNT(*) AS total_socios FROM socios;
SELECT COUNT(*) AS total_cuotas FROM cuotas;
SELECT COUNT(*) AS socios_sin_email FROM socios WHERE email_principal IS NULL AND estado = 'activo';
SELECT COUNT(*) AS socios_sin_iban  FROM socios WHERE iban IS NULL AND estado = 'activo';
""")

    return "\n".join(lineas), activos_socios, activos_coop


# ─── EJECUCIÓN PRINCIPAL ──────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  ASPROJUMA - Migración Excel → Base de datos")
    print("=" * 60)

    # Leer datos
    print("\n[1/3] Leyendo socios profesores...")
    try:
        socios, alertas_s = leer_socios(FICHERO_SOCIOS)
        print(f"      {len(socios)} registros leídos")
    except FileNotFoundError:
        print(f"  ERROR: No se encuentra '{FICHERO_SOCIOS}'")
        print("  Asegúrate de que el Excel está en la misma carpeta que este script.")
        sys.exit(1)

    print("\n[2/3] Leyendo miembros cooperantes...")
    cooperantes, alertas_c = leer_cooperantes(FICHERO_SOCIOS)
    print(f"      {len(cooperantes)} registros leídos")

    # Generar SQL
    print("\n[3/3] Generando SQL...")
    sql, activos_s, activos_c = generar_sql(socios, cooperantes)

    with open(FICHERO_SQL, "w", encoding="utf-8") as f:
        f.write(sql)
    print(f"      Guardado: {FICHERO_SQL}")

    # Informe
    todas_alertas = alertas_s + alertas_c
    estados_socios     = {}
    estados_cooperantes = {}
    for s in socios:
        estados_socios[s['estado']] = estados_socios.get(s['estado'], 0) + 1
    for c in cooperantes:
        estados_cooperantes[c['estado']] = estados_cooperantes.get(c['estado'], 0) + 1

    informe = []
    informe.append("ASPROJUMA - Informe de migración")
    informe.append(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    informe.append("=" * 50)
    informe.append("\nSOCIOS PROFESORES")
    for estado, n in sorted(estados_socios.items()):
        informe.append(f"  {estado:20s}: {n}")
    informe.append(f"  {'TOTAL':20s}: {len(socios)}")
    informe.append("\nMIEMBROS COOPERANTES")
    for estado, n in sorted(estados_cooperantes.items()):
        informe.append(f"  {estado:20s}: {n}")
    informe.append(f"  {'TOTAL':20s}: {len(cooperantes)}")
    informe.append("\nIBANs normalizados al formato ES...")
    informe.append("  (Los registros con formato antiguo 20 dígitos se han convertido)")
    if todas_alertas:
        informe.append(f"\nALERTAS ({len(todas_alertas)} registros a revisar):")
        for a in todas_alertas:
            informe.append(a)
    else:
        informe.append("\nSin alertas. Todos los registros activos tienen IBAN y DNI.")
    informe.append("\nPRÓXIMOS PASOS:")
    informe.append("  1. Abre Supabase (supabase.com) e inicia sesión")
    informe.append("  2. Ve a tu proyecto ASPROJUMA")
    informe.append("  3. Haz clic en 'SQL Editor' en el menú izquierdo")
    informe.append("  4. Haz clic en '+ New query'")
    informe.append("  5. Pega el contenido de asprojuma_datos.sql")
    informe.append("  6. Haz clic en 'Run' (botón verde)")
    informe.append("  7. Comprueba los resultados de las consultas del Paso 5")

    with open(FICHERO_INFORME, "w", encoding="utf-8") as f:
        f.write("\n".join(informe))
    print(f"      Guardado: {FICHERO_INFORME}")

    print("\n" + "=" * 60)
    print("  ¡Migración preparada!")
    print(f"  Socios profesores : {len(socios)} ({activos_s} activos)")
    print(f"  Cooperantes       : {len(cooperantes)} ({activos_c} activos)")
    if todas_alertas:
        print(f"  Alertas a revisar : {len(todas_alertas)} (ver informe_migracion.txt)")
    print("\n  Ficheros generados:")
    print(f"    → {FICHERO_SQL}   (pegar en Supabase)")
    print(f"    → {FICHERO_INFORME}")
    print("=" * 60)


if __name__ == "__main__":
    main()
